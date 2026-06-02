"""
Run once: adds 'Estado' column to CRM.xlsx.
Then upload to Google Drive and open with Google Sheets.
"""
import openpyxl
from pathlib import Path

CRM_PATH = Path(r"C:\Users\dani\Desktop\Proyectos\CRM\CRM.xlsx")

wb = openpyxl.load_workbook(CRM_PATH)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print("Columnas actuales:", headers)

if 'Estado' not in headers:
    col = len(headers) + 1
    ws.cell(row=1, column=col, value='Estado')
    count = 0
    for row in ws.iter_rows(min_row=2):
        if any(cell.value for cell in row):
            row[col - 1].value = 'Sin contactar'
            count += 1
    wb.save(CRM_PATH)
    print(f"Columna 'Estado' añadida. {count} filas marcadas 'Sin contactar'. Guardado en {CRM_PATH}")
else:
    print("La columna 'Estado' ya existe.")
